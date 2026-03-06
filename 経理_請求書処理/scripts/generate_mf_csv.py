"""
仕訳データからMoneyForwardインポート用CSVを生成するスクリプト
ダウンロードしたExcel/CSVデータを読み込み、法人別にMFインポートCSVを出力する

使い方:
  python generate_mf_csv.py --input-dir data/ --output-dir output/
  python generate_mf_csv.py --file data/恵聖会/仕訳データ.xlsx --entity 恵聖会
  python generate_mf_csv.py --file data/ルミナス/仕訳データ.csv --entity ルミナス
"""
import os
import sys
import csv
import json
import argparse
import re
from datetime import datetime

# 同じディレクトリのモジュールをインポート
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from account_mapping_rules import (
    KEYWORD_RULES,
    ENTITY_MAP,
    classify_transaction,
    get_credit_account,
    get_entity,
    create_journal_entry,
)


def detect_encoding(filepath):
    """ファイルのエンコーディングを検出"""
    encodings = ["utf-8", "shift-jis", "cp932", "euc-jp", "iso-2022-jp"]
    for enc in encodings:
        try:
            with open(filepath, "r", encoding=enc) as f:
                f.read(1000)
            return enc
        except (UnicodeDecodeError, UnicodeError):
            continue
    return "utf-8"


def read_excel_file(filepath):
    """Excelファイルから仕訳データを読み込み"""
    from openpyxl import load_workbook

    wb = load_workbook(filepath, data_only=True)
    ws = wb.active

    # ヘッダー行を検出
    headers = []
    header_row = 1
    for row in ws.iter_rows(min_row=1, max_row=5, values_only=False):
        cells = [cell.value for cell in row]
        if any(
            h in str(c or "")
            for c in cells
            for h in ["日付", "取引日", "借方", "勘定科目", "摘要"]
        ):
            headers = [str(c or "").strip() for c in cells]
            header_row = row[0].row
            break

    if not headers:
        # ヘッダーが見つからない場合、1行目をヘッダーとする
        first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = [str(c or "").strip() for c in first_row]
        header_row = 1

    # データ行を読み込み
    entries = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if all(c is None for c in row):
            continue
        if row[0] and str(row[0]).strip() in ("合計", "計", ""):
            continue

        entry = {}
        for i, header in enumerate(headers):
            if i < len(row):
                entry[header] = row[i]
        entries.append(entry)

    wb.close()
    return headers, entries


def read_csv_file(filepath):
    """CSVファイルから仕訳データを読み込み"""
    encoding = detect_encoding(filepath)
    entries = []
    headers = []

    # TSVかCSVか検出
    with open(filepath, "r", encoding=encoding) as f:
        first_line = f.readline()
        delimiter = "\t" if "\t" in first_line else ","

    with open(filepath, "r", encoding=encoding) as f:
        reader = csv.reader(f, delimiter=delimiter)
        headers = next(reader, [])
        headers = [h.strip() for h in headers]
        for row in reader:
            if not any(row):
                continue
            entry = {}
            for i, header in enumerate(headers):
                if i < len(row):
                    entry[header] = row[i]
            entries.append(entry)

    return headers, entries


def normalize_entry(raw_entry, headers):
    """
    様々な形式の入力データを統一フォーマットに正規化
    対応形式:
    - 仕訳一覧形式（日付, 借方勘定科目, 借方金額, 貸方勘定科目, 貸方金額, 税区分, 摘要）
    - 銀行取引形式（日付, 摘要, 出金, 入金, 残高）
    - MFエクスポート形式（取引日, 借方勘定科目, ...）
    """
    # カラム名のマッピング（様々な形式に対応）
    date_keys = ["日付", "取引日", "date", "Date"]
    debit_acc_keys = ["借方勘定科目", "借方科目", "debit_account"]
    debit_amt_keys = ["借方金額", "debit_amount", "出金", "お引出し", "引出金額"]
    credit_acc_keys = ["貸方勘定科目", "貸方科目", "credit_account"]
    credit_amt_keys = ["貸方金額", "credit_amount", "入金", "お預入", "預入金額"]
    tax_keys = ["税区分", "tax_category"]
    desc_keys = ["摘要", "description", "備考", "お取引内容", "取引内容"]
    vendor_keys = ["取引先", "vendor"]
    status_keys = ["判定", "status"]
    account_keys = ["口座", "口座名", "bank_account"]

    def find_value(entry, key_list):
        for k in key_list:
            if k in entry and entry[k] is not None:
                return entry[k]
        return None

    date = find_value(raw_entry, date_keys) or ""
    debit_acc = find_value(raw_entry, debit_acc_keys) or ""
    debit_amt = find_value(raw_entry, debit_amt_keys) or 0
    credit_acc = find_value(raw_entry, credit_acc_keys) or ""
    credit_amt = find_value(raw_entry, credit_amt_keys) or 0
    tax = find_value(raw_entry, tax_keys) or ""
    desc = find_value(raw_entry, desc_keys) or ""
    vendor = find_value(raw_entry, vendor_keys) or ""
    status = find_value(raw_entry, status_keys) or "自動"
    bank_account = find_value(raw_entry, account_keys) or ""

    # 日付の正規化
    if isinstance(date, datetime):
        date = date.strftime("%Y/%m/%d")
    else:
        date = str(date).strip()
        # YYYY-MM-DD → YYYY/MM/DD
        date = date.replace("-", "/")

    # 金額の正規化
    def parse_amount(val):
        if val is None or val == "":
            return 0
        if isinstance(val, (int, float)):
            return abs(int(val))
        val = str(val).replace(",", "").replace("¥", "").replace("￥", "").strip()
        try:
            return abs(int(float(val)))
        except ValueError:
            return 0

    debit_amt = parse_amount(debit_amt)
    credit_amt = parse_amount(credit_amt)

    # 銀行取引形式の場合（借方/貸方科目がない場合は自動分類）
    if not debit_acc and not credit_acc and desc:
        is_deposit = credit_amt > 0 and debit_amt == 0
        amount = credit_amt if is_deposit else debit_amt

        if bank_account:
            entry = create_journal_entry(date, str(desc), amount, bank_account, is_deposit)
            return entry

        # 口座不明でも科目は判定
        account, tax_cat, auto_status = classify_transaction(str(desc))
        if is_deposit:
            debit_acc = "普通預金"
            credit_acc = account if account != "不明" else "売上高"
        else:
            debit_acc = account
            credit_acc = "普通預金"
        tax = tax_cat
        status = auto_status

    # 金額の整合
    amount = debit_amt or credit_amt
    if debit_amt == 0:
        debit_amt = credit_amt
    if credit_amt == 0:
        credit_amt = debit_amt

    return {
        "date": date,
        "debit_account": str(debit_acc).strip(),
        "debit_amount": debit_amt,
        "credit_account": str(credit_acc).strip(),
        "credit_amount": credit_amt,
        "tax_category": str(tax).strip(),
        "description": str(desc).strip(),
        "vendor": str(vendor).strip(),
        "invoice_number": "",
        "status": str(status).strip(),
        "entity": get_entity(str(bank_account)) if bank_account else "",
        "bank_account": str(bank_account).strip(),
    }


def write_mf_import_csv(entries, output_path):
    """MoneyForwardインポート用CSV出力（Shift-JIS）"""
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)

    with open(output_path, "w", encoding="shift-jis", newline="", errors="replace") as f:
        writer = csv.writer(f)
        # MFクラウド会計の仕訳インポート形式
        writer.writerow([
            "取引日",
            "借方勘定科目",
            "借方補助科目",
            "借方税区分",
            "借方金額",
            "貸方勘定科目",
            "貸方補助科目",
            "貸方税区分",
            "貸方金額",
            "摘要",
        ])
        for entry in entries:
            if not entry.get("date") or not entry.get("debit_account"):
                continue
            writer.writerow([
                entry["date"],
                entry["debit_account"],
                "",  # 借方補助科目
                entry.get("tax_category", ""),
                entry["debit_amount"],
                entry["credit_account"],
                "",  # 貸方補助科目
                "",  # 貸方税区分
                entry["credit_amount"],
                entry.get("description", ""),
            ])

    return output_path


def write_journal_excel(entries, output_path):
    """仕訳一覧Excel出力"""
    from create_journal_excel import create_journal_excel
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    create_journal_excel(entries, output_path)
    return output_path


def write_journal_json(entries, output_path):
    """仕訳データJSON出力（中間ファイル）"""
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(entries, f, ensure_ascii=False, indent=2)
    return output_path


def process_file(filepath):
    """1ファイルを読み込んで正規化した仕訳エントリのリストを返す"""
    ext = os.path.splitext(filepath)[1].lower()

    if ext in (".xlsx", ".xls"):
        headers, raw_entries = read_excel_file(filepath)
    elif ext in (".csv", ".tsv", ".txt"):
        headers, raw_entries = read_csv_file(filepath)
    else:
        print(f"  ⚠ 未対応ファイル形式: {filepath}")
        return []

    entries = []
    for raw in raw_entries:
        entry = normalize_entry(raw, headers)
        if entry and entry.get("date"):
            entries.append(entry)

    return entries


def main():
    parser = argparse.ArgumentParser(description="仕訳データからMFインポートCSVを生成")
    parser.add_argument("--input-dir", help="入力データディレクトリ")
    parser.add_argument("--file", help="個別入力ファイル")
    parser.add_argument("--entity", choices=["ルミナス", "恵聖会", "both"], default="both")
    parser.add_argument(
        "--output-dir",
        default=os.path.join(
            os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "output"
        ),
    )
    parser.add_argument("--month", help="対象月（例: 2026-03）")
    args = parser.parse_args()

    all_entries = []
    month_suffix = args.month.replace("-", "") if args.month else datetime.now().strftime("%Y%m")

    # ファイル読み込み
    if args.file:
        print(f"ファイル読み込み: {args.file}")
        entries = process_file(args.file)
        print(f"  → {len(entries)}件")
        all_entries.extend(entries)
    elif args.input_dir:
        print(f"ディレクトリ読み込み: {args.input_dir}")
        for root, dirs, files in os.walk(args.input_dir):
            for f in sorted(files):
                if f.startswith("~$") or f.startswith("."):
                    continue
                ext = os.path.splitext(f)[1].lower()
                if ext in (".xlsx", ".xls", ".csv", ".tsv", ".txt"):
                    filepath = os.path.join(root, f)
                    print(f"  読み込み: {filepath}")
                    entries = process_file(filepath)
                    print(f"    → {len(entries)}件")
                    all_entries.extend(entries)
    else:
        print("エラー: --input-dir または --file を指定してください")
        sys.exit(1)

    if not all_entries:
        print("\n⚠ 仕訳データが見つかりませんでした")
        sys.exit(1)

    # 法人別に分割
    luminous_entries = []
    keiseikai_entries = []
    unknown_entries = []

    for entry in all_entries:
        entity = entry.get("entity", "")
        if entity == "ルミナス":
            luminous_entries.append(entry)
        elif entity == "恵聖会":
            keiseikai_entries.append(entry)
        else:
            unknown_entries.append(entry)

    # サマリー表示
    print(f"\n{'='*60}")
    print(f"処理結果サマリー")
    print(f"{'='*60}")
    print(f"  全件数: {len(all_entries)}")
    print(f"  ルミナス: {len(luminous_entries)}件")
    print(f"  恵聖会: {len(keiseikai_entries)}件")
    if unknown_entries:
        print(f"  法人不明: {len(unknown_entries)}件")

    # 自動/要確認の内訳
    auto_count = sum(1 for e in all_entries if e.get("status") == "自動")
    review_count = sum(1 for e in all_entries if e.get("status") == "要確認")
    print(f"  自動判定: {auto_count}件 ({auto_count*100//max(len(all_entries),1)}%)")
    print(f"  要確認: {review_count}件")

    # 出力
    output_base = args.output_dir
    os.makedirs(output_base, exist_ok=True)

    def output_entity(entity_name, entries):
        if not entries:
            return
        entity_dir = os.path.join(output_base, entity_name)
        os.makedirs(entity_dir, exist_ok=True)

        # MFインポートCSV
        csv_path = os.path.join(entity_dir, f"MFインポート_{entity_name}_{month_suffix}.csv")
        write_mf_import_csv(entries, csv_path)
        print(f"\n✓ MF CSV: {csv_path}")

        # 仕訳一覧Excel
        xlsx_path = os.path.join(entity_dir, f"仕訳一覧_{entity_name}_{month_suffix}.xlsx")
        try:
            write_journal_excel(entries, xlsx_path)
            print(f"✓ Excel: {xlsx_path}")
        except Exception as e:
            print(f"⚠ Excel出力エラー: {e}")

        # JSON（中間データ）
        json_path = os.path.join(entity_dir, f"仕訳データ_{entity_name}_{month_suffix}.json")
        write_journal_json(entries, json_path)
        print(f"✓ JSON: {json_path}")

        # 要確認リスト
        review = [e for e in entries if e.get("status") == "要確認"]
        if review:
            review_path = os.path.join(entity_dir, f"要確認_{entity_name}_{month_suffix}.csv")
            write_mf_import_csv(review, review_path)
            print(f"⚠ 要確認 {len(review)}件: {review_path}")

    if args.entity in ("ルミナス", "both"):
        output_entity("ルミナス", luminous_entries)
    if args.entity in ("恵聖会", "both"):
        output_entity("恵聖会", keiseikai_entries)

    # 法人不明分も出力
    if unknown_entries:
        output_entity("不明", unknown_entries)

    print(f"\n{'='*60}")
    print(f"完了！ 出力先: {output_base}")
    print(f"{'='*60}")
    print(f"\n次のステップ:")
    print(f"  1. 要確認データを手動で確認・修正")
    print(f"  2. MFインポートCSVをMoneyForwardにアップロード")
    print(f"     MF → 会計帳簿 → 仕訳帳 → インポート")


if __name__ == "__main__":
    main()
