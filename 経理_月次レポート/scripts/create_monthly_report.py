"""
月次経理レポート作成スクリプト
仕訳データから資金繰り表・試算表・月次推移表をExcelで出力
使い方: python create_monthly_report.py --data journal_data.json --month 2026-03 --output 月次レポート_202603.xlsx
"""
import json
import sys
from datetime import datetime

def create_monthly_report(data, month, output_path, prev_balance=0):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # === シート1: 資金繰り表 ===
    ws1 = wb.active
    ws1.title = "資金繰り表"

    title_font = Font(bold=True, size=16, name='Arial')
    section_font = Font(bold=True, size=12, name='Arial')
    section_fill = PatternFill('solid', fgColor='D6E4F0')
    total_font = Font(bold=True, size=11, name='Arial')
    money_fmt = '¥#,##0;[Red](¥#,##0);-'
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    top_border = Border(top=Side(style='double'))

    # タイトル
    ws1['A1'] = f'医療法人 恵聖会 資金繰り表'
    ws1['A1'].font = title_font
    ws1.merge_cells('A1:B1')
    ws1['A2'] = f'対象月: {month}'
    ws1['A2'].font = Font(size=11, name='Arial')

    # 収入カテゴリ
    income_categories = ['売上高', '診療報酬（社保）', '診療報酬（国保）', '窓口収入', '自費診療', 'その他収入']
    # 支出カテゴリ（人件費）
    personnel_categories = ['給料賃金', '法定福利費']
    # 支出カテゴリ（経費）
    expense_categories = [
        '地代家賃', '水道光熱費', '通信費', '医薬品費', '備品・消耗品費',
        '業務委託料', 'リース料', '支払報酬', '広告宣伝費', '修繕費',
        '保険料', '接待交際費', '旅費交通費', '研修費', '支払手数料', 'その他経費'
    ]

    # 集計
    income_totals = {}
    expense_totals = {}
    for entry in data:
        acct = entry.get('debit_account', '') or entry.get('credit_account', '')
        amount = entry.get('debit_amount', 0) or entry.get('credit_amount', 0)
        if acct in income_categories:
            income_totals[acct] = income_totals.get(acct, 0) + amount
        elif acct in personnel_categories + expense_categories:
            expense_totals[acct] = expense_totals.get(acct, 0) + amount

    row = 4
    # 収入の部
    ws1.cell(row=row, column=1, value='【収入の部】').font = section_font
    ws1.cell(row=row, column=1).fill = section_fill
    ws1.cell(row=row, column=2).fill = section_fill
    row += 1
    income_start = row
    for cat in income_categories:
        ws1.cell(row=row, column=1, value=f'  {cat}')
        ws1.cell(row=row, column=2, value=income_totals.get(cat, 0)).number_format = money_fmt
        row += 1
    income_end = row - 1
    ws1.cell(row=row, column=1, value='  収入合計').font = total_font
    ws1.cell(row=row, column=2, value=f'=SUM(B{income_start}:B{income_end})').font = total_font
    ws1.cell(row=row, column=2).number_format = money_fmt
    ws1.cell(row=row, column=2).border = top_border
    income_total_row = row
    row += 2

    # 支出の部 - 人件費
    ws1.cell(row=row, column=1, value='【支出の部】').font = section_font
    ws1.cell(row=row, column=1).fill = section_fill
    ws1.cell(row=row, column=2).fill = section_fill
    row += 1
    ws1.cell(row=row, column=1, value='  ＜人件費＞').font = Font(bold=True, name='Arial')
    row += 1
    personnel_start = row
    for cat in personnel_categories:
        ws1.cell(row=row, column=1, value=f'    {cat}')
        ws1.cell(row=row, column=2, value=expense_totals.get(cat, 0)).number_format = money_fmt
        row += 1
    personnel_end = row - 1
    ws1.cell(row=row, column=1, value='  人件費小計').font = total_font
    ws1.cell(row=row, column=2, value=f'=SUM(B{personnel_start}:B{personnel_end})').font = total_font
    ws1.cell(row=row, column=2).number_format = money_fmt
    personnel_total_row = row
    row += 1

    # 支出の部 - 経費
    ws1.cell(row=row, column=1, value='  ＜経費＞').font = Font(bold=True, name='Arial')
    row += 1
    expense_start = row
    for cat in expense_categories:
        ws1.cell(row=row, column=1, value=f'    {cat}')
        ws1.cell(row=row, column=2, value=expense_totals.get(cat, 0)).number_format = money_fmt
        row += 1
    expense_end = row - 1
    ws1.cell(row=row, column=1, value='  経費小計').font = total_font
    ws1.cell(row=row, column=2, value=f'=SUM(B{expense_start}:B{expense_end})').font = total_font
    ws1.cell(row=row, column=2).number_format = money_fmt
    expense_total_row = row
    row += 1

    ws1.cell(row=row, column=1, value='  支出合計').font = total_font
    ws1.cell(row=row, column=2, value=f'=B{personnel_total_row}+B{expense_total_row}').font = total_font
    ws1.cell(row=row, column=2).number_format = money_fmt
    ws1.cell(row=row, column=2).border = top_border
    expense_total_sum_row = row
    row += 2

    # 差引
    ws1.cell(row=row, column=1, value='【差引】').font = section_font
    ws1.cell(row=row, column=1).fill = PatternFill('solid', fgColor='E2EFDA')
    ws1.cell(row=row, column=2).fill = PatternFill('solid', fgColor='E2EFDA')
    row += 1
    ws1.cell(row=row, column=1, value='  当月収支').font = total_font
    ws1.cell(row=row, column=2, value=f'=B{income_total_row}-B{expense_total_sum_row}').font = total_font
    ws1.cell(row=row, column=2).number_format = money_fmt
    net_row = row
    row += 1
    ws1.cell(row=row, column=1, value='  前月繰越残高')
    ws1.cell(row=row, column=2, value=prev_balance).number_format = money_fmt
    prev_row = row
    row += 1
    ws1.cell(row=row, column=1, value='  翌月繰越残高').font = Font(bold=True, size=12, name='Arial')
    ws1.cell(row=row, column=2, value=f'=B{net_row}+B{prev_row}')
    ws1.cell(row=row, column=2).font = Font(bold=True, size=12, name='Arial')
    ws1.cell(row=row, column=2).number_format = money_fmt

    ws1.column_dimensions['A'].width = 30
    ws1.column_dimensions['B'].width = 18

    # === シート2: 試算表 ===
    ws2 = wb.create_sheet('試算表')
    ws2['A1'] = '医療法人 恵聖会 残高試算表'
    ws2['A1'].font = title_font
    ws2.merge_cells('A1:C1')
    ws2['A2'] = f'{month} 現在'

    headers2 = ['勘定科目', '借方残高', '貸方残高']
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=4, column=col, value=h)
        cell.font = Font(bold=True, name='Arial')
        cell.fill = PatternFill('solid', fgColor='D9E1F2')
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    debit_accounts = {
        '【資産の部】': None,
        '  現金': 'debit', '  普通預金（SMBC）': 'debit', '  普通預金（常陽）': 'debit',
        '  普通預金（楽天）': 'debit', '  売掛金': 'debit', '  未収金': 'debit',
        '【費用の部】': None,
    }
    for cat in personnel_categories + expense_categories:
        debit_accounts[f'  {cat}'] = 'debit'

    credit_accounts = {
        '【負債の部】': None,
        '  未払金': 'credit', '  預り金': 'credit', '  借入金': 'credit',
        '【純資産の部】': None,
        '  元入金': 'credit',
        '【収益の部】': None,
    }
    for cat in income_categories:
        credit_accounts[f'  {cat}'] = 'credit'

    row2 = 5
    all_accounts = {**debit_accounts, **credit_accounts}
    account_balances = {}
    for entry in data:
        d_acct = entry.get('debit_account', '')
        c_acct = entry.get('credit_account', '')
        d_amt = entry.get('debit_amount', 0) or 0
        c_amt = entry.get('credit_amount', 0) or 0
        if d_acct:
            account_balances[d_acct] = account_balances.get(d_acct, {'debit': 0, 'credit': 0})
            account_balances[d_acct]['debit'] += d_amt
        if c_acct:
            account_balances[c_acct] = account_balances.get(c_acct, {'debit': 0, 'credit': 0})
            account_balances[c_acct]['credit'] += c_amt

    debit_start = row2
    for acct, side in all_accounts.items():
        ws2.cell(row=row2, column=1, value=acct).border = thin_border
        if side is None:
            ws2.cell(row=row2, column=1).font = Font(bold=True, name='Arial')
            ws2.cell(row=row2, column=1).fill = PatternFill('solid', fgColor='F2F2F2')
        else:
            clean_name = acct.strip()
            bal = account_balances.get(clean_name, {'debit': 0, 'credit': 0})
            if side == 'debit':
                ws2.cell(row=row2, column=2, value=bal['debit']).number_format = money_fmt
            else:
                ws2.cell(row=row2, column=3, value=bal['credit']).number_format = money_fmt
        ws2.cell(row=row2, column=2).border = thin_border
        ws2.cell(row=row2, column=3).border = thin_border
        row2 += 1

    ws2.cell(row=row2, column=1, value='合計').font = Font(bold=True, size=12, name='Arial')
    ws2.cell(row=row2, column=2, value=f'=SUM(B{debit_start}:B{row2-1})').font = total_font
    ws2.cell(row=row2, column=2).number_format = money_fmt
    ws2.cell(row=row2, column=3, value=f'=SUM(C{debit_start}:C{row2-1})').font = total_font
    ws2.cell(row=row2, column=3).number_format = money_fmt

    ws2.column_dimensions['A'].width = 28
    ws2.column_dimensions['B'].width = 16
    ws2.column_dimensions['C'].width = 16

    wb.save(output_path)
    return output_path


if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument('--data', required=True)
    parser.add_argument('--month', required=True)
    parser.add_argument('--output', default='月次レポート.xlsx')
    parser.add_argument('--prev-balance', type=int, default=0)
    args = parser.parse_args()

    with open(args.data, 'r', encoding='utf-8') as f:
        data = json.load(f)

    create_monthly_report(data, args.month, args.output, args.prev_balance)
    print(f"Output: {args.output}")
