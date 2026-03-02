"""
請求書データからExcel仕訳一覧とMFインポートCSVを作成するスクリプト
使い方: python create_journal_excel.py --data journal_data.json --output 仕訳一覧.xlsx --csv mf_import.csv
"""
import json, sys, os, csv
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def create_journal_excel(data, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "仕訳一覧"
    header_font = Font(bold=True, size=11, name='Arial')
    header_fill = PatternFill('solid', fgColor='D9E1F2')
    warning_fill = PatternFill('solid', fgColor='FFFF00')
    money_format = '#,##0'
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    headers = ['日付', '借方勘定科目', '借方金額', '貸方勘定科目', '貸方金額', '税区分', '摘要', '取引先', '請求書番号', '判定']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    for i, entry in enumerate(data, 2):
        ws.cell(row=i, column=1, value=entry.get('date', '')).border = thin_border
        ws.cell(row=i, column=2, value=entry.get('debit_account', '')).border = thin_border
        c3 = ws.cell(row=i, column=3, value=entry.get('debit_amount', 0))
        c3.number_format = money_format
        c3.border = thin_border
        ws.cell(row=i, column=4, value=entry.get('credit_account', '')).border = thin_border
        c5 = ws.cell(row=i, column=5, value=entry.get('credit_amount', 0))
        c5.number_format = money_format
        c5.border = thin_border
        ws.cell(row=i, column=6, value=entry.get('tax_category', '')).border = thin_border
        ws.cell(row=i, column=7, value=entry.get('description', '')).border = thin_border
        ws.cell(row=i, column=8, value=entry.get('vendor', '')).border = thin_border
        ws.cell(row=i, column=9, value=entry.get('invoice_number', '')).border = thin_border
        status_cell = ws.cell(row=i, column=10, value=entry.get('status', '自動'))
        status_cell.border = thin_border
        if entry.get('status') == '要確認':
            for col in range(1, 11):
                ws.cell(row=i, column=col).fill = warning_fill

    widths = [12, 16, 14, 16, 14, 16, 30, 20, 14, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    last_row = len(data) + 1
    total_row = last_row + 1
    ws.cell(row=total_row, column=2, value='合計').font = Font(bold=True)
    ws.cell(row=total_row, column=3, value=f'=SUM(C2:C{last_row})').font = Font(bold=True)
    ws.cell(row=total_row, column=3).number_format = money_format
    ws.cell(row=total_row, column=5, value=f'=SUM(E2:E{last_row})').font = Font(bold=True)
    ws.cell(row=total_row, column=5).number_format = money_format
    wb.save(output_path)

def create_mf_csv(data, csv_path):
    with open(csv_path, 'w', encoding='shift-jis', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(['取引日', '借方勘定科目', '借方金額', '貸方勘定科目', '貸方金額', '摘要'])
        for entry in data:
            writer.writerow([entry.get('date',''), entry.get('debit_account',''), entry.get('debit_amount',0), entry.get('credit_account',''), entry.get('credit_amount',0), entry.get('description','')])

if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument('--data', required=True)
    parser.add_argument('--output', default='仕訳一覧.xlsx')
    parser.add_argument('--csv', default='mf_import.csv')
    args = parser.parse_args()
    with open(args.data, 'r', encoding='utf-8') as f:
        data = json.load(f)
    create_journal_excel(data, args.output)
    create_mf_csv(data, args.csv)
    print(f"Excel: {args.output}")
    print(f"CSV: {args.csv}")
