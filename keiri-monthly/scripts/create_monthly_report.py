"""
月次経理レポート作成スクリプト
仕訳データから資金繰り表・試算表をExcelで出力
使い方: python create_monthly_report.py --data journal_data.json --month 2026-03 --output 月次レポート.xlsx
"""
import json, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def create_monthly_report(data, month, output_path, prev_balance=0):
    wb = Workbook()
    title_font = Font(bold=True, size=16, name='Arial')
    section_font = Font(bold=True, size=12, name='Arial')
    section_fill = PatternFill('solid', fgColor='D6E4F0')
    total_font = Font(bold=True, size=11, name='Arial')
    money_fmt = '#,##0;[Red](#,##0);-'
    top_border = Border(top=Side(style='double'))

    income_categories = ['売上高', '診療報酬（社保）', '診療報酬（国保）', '窓口収入', '自費診療', 'その他収入']
    personnel_categories = ['給料賃金', '法定福利費']
    expense_categories = ['地代家賃', '水道光熱費', '通信費', '医薬品費', '備品・消耗品費', '業務委託料', 'リース料', '支払報酬', '広告宣伝費', '修繕費', '保険料', '接待交際費', '旅費交通費', '研修費', '支払手数料', 'その他経費']

    income_totals = {}
    expense_totals = {}
    for entry in data:
        acct = entry.get('debit_account', '') or entry.get('credit_account', '')
        amount = entry.get('debit_amount', 0) or entry.get('credit_amount', 0)
        if acct in income_categories:
            income_totals[acct] = income_totals.get(acct, 0) + amount
        elif acct in personnel_categories + expense_categories:
            expense_totals[acct] = expense_totals.get(acct, 0) + amount

    # シート1: 資金繰り表
    ws1 = wb.active
    ws1.title = "資金繰り表"
    ws1['A1'] = f'医療法人 恵聖会 資金繰り表'
    ws1['A1'].font = title_font
    ws1.merge_cells('A1:B1')
    ws1['A2'] = f'対象月: {month}'

    row = 4
    ws1.cell(row=row, column=1, value='【収入の部】').font = section_font
    ws1.cell(row=row, column=1).fill = section_fill
    ws1.cell(row=row, column=2).fill = section_fill
    row += 1
    income_start = row
    for cat in income_categories:
        ws1.cell(row=row, column=1, value=f'  {cat}')
        ws1.cell(row=row, column=2, value=income_totals.get(cat, 0)).number_format = money_fmt
        row += 1
    income_end = row - 1
    ws1.cell(row=row, column=1, value='  収入合計').font = total_font
    ws1.cell(row=row, column=2, value=f'=SUM(B{income_start}:B{income_end})').font = total_font
    ws1.cell(row=row, column=2).number_format = money_fmt
    income_total_row = row
    row += 2

    ws1.cell(row=row, column=1, value='【支出の部】').font = section_font
    ws1.cell(row=row, column=1).fill = section_fill
    ws1.cell(row=row, column=2).fill = section_fill
    row += 1
    personnel_start = row
    for cat in personnel_categories:
        ws1.cell(row=row, column=1, value=f'  {cat}')
        ws1.cell(row=row, column=2, value=expense_totals.get(cat, 0)).number_format = money_fmt
        row += 1
    personnel_end = row - 1
    ws1.cell(row=row, column=1, value='  人件費小計').font = total_font
    ws1.cell(row=row, column=2, value=f'=SUM(B{personnel_start}:B{personnel_end})').font = total_font
    ws1.cell(row=row, column=2).number_format = money_fmt
    personnel_total_row = row
    row += 1
    expense_start = row
    for cat in expense_categories:
        ws1.cell(row=row, column=1, value=f'  {cat}')
        ws1.cell(row=row, column=2, value=expense_totals.get(cat, 0)).number_format = money_fmt
        row += 1
    expense_end = row - 1
    ws1.cell(row=row, column=1, value='  経費小計').font = total_font
    ws1.cell(row=row, column=2, value=f'=SUM(B{expense_start}:B{expense_end})').font = total_font
    ws1.cell(row=row, column=2).number_format = money_fmt
    expense_total_row = row
    row += 1
    ws1.cell(row=row, column=1, value='  支出合計').font = total_font
    ws1.cell(row=row, column=2, value=f'=B{personnel_total_row}+B{expense_total_row}').font = total_font
    ws1.cell(row=row, column=2).number_format = money_fmt
    expense_sum_row = row
    row += 2

    ws1.cell(row=row, column=1, value='【差引】').font = section_font
    ws1.cell(row=row, column=1).fill = PatternFill('solid', fgColor='E2EFDA')
    ws1.cell(row=row, column=2).fill = PatternFill('solid', fgColor='E2EFDA')
    row += 1
    ws1.cell(row=row, column=1, value='  当月収支').font = total_font
    ws1.cell(row=row, column=2, value=f'=B{income_total_row}-B{expense_sum_row}').font = total_font
    ws1.cell(row=row, column=2).number_format = money_fmt
    net_row = row
    row += 1
    ws1.cell(row=row, column=1, value='  前月繰越残高')
    ws1.cell(row=row, column=2, value=prev_balance).number_format = money_fmt
    prev_row = row
    row += 1
    ws1.cell(row=row, column=1, value='  翌月繰越残高').font = Font(bold=True, size=12)
    ws1.cell(row=row, column=2, value=f'=B{net_row}+B{prev_row}').font = Font(bold=True, size=12)
    ws1.cell(row=row, column=2).number_format = money_fmt
    ws1.column_dimensions['A'].width = 30
    ws1.column_dimensions['B'].width = 18

    # シート2: 試算表
    ws2 = wb.create_sheet('試算表')
    ws2['A1'] = '医療法人 恵聖会 残高試算表'
    ws2['A1'].font = title_font
    ws2.merge_cells('A1:C1')
    ws2['A2'] = f'{month} 現在'
    for col, h in enumerate(['勘定科目', '借方残高', '貸方残高'], 1):
        cell = ws2.cell(row=4, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill('solid', fgColor='D9E1F2')
    ws2.column_dimensions['A'].width = 28
    ws2.column_dimensions['B'].width = 16
    ws2.column_dimensions['C'].width = 16

    wb.save(output_path)
    return output_path

if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument('--data', required=True)
    parser.add_argument('--month', required=True)
    parser.add_argument('--output', default='月次レポート.xlsx')
    parser.add_argument('--prev-balance', type=int, default=0)
    args = parser.parse_args()
    with open(args.data, 'r', encoding='utf-8') as f:
        data = json.load(f)
    create_monthly_report(data, args.month, args.output, args.prev_balance)
    print(f"Output: {args.output}")
